import os
import pandas as pd
import matplotlib.pyplot as plt
import glob
import openpyxl
import time
import seaborn as sns
import numpy as np

# Set the font sizes for all plots
plt.rcParams.update(
    {
        "font.size": 12,  # General font size
        "axes.labelsize": 14,  # Axis labels
        "axes.titlesize": 14,  # Title size reduced
        "xtick.labelsize": 12,  # X-axis tick labels
        "ytick.labelsize": 12,  # Y-axis tick labels
        "legend.fontsize": 12,  # Legend font size
        "figure.titlesize": 14,  # Figure title size reduced
        "figure.figsize": (12, 8),  # Default figure size
    }
)


###########-------UNIT CONVERSION-------###########
def mwh_to_bcm(mwh, energy_content_mj_per_m3=35.8):
    return mwh * 3.6 * 1000 / (energy_content_mj_per_m3 * 1e9)


###################################################


def plot_nzia_benchmark(output_file_path):
    # Extract scenario name from file name
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")

    # Create output folder next to the Excel file
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    component_sheets = [
        "capacity_ext_stockout",
        "capacity_ext_euprimary",
        "capacity_ext_eusecondary",
        "capacity_ext_imported"  # Add imported to read the data
    ]

    # Show actual sheet names
    wb = openpyxl.load_workbook(output_file_path, read_only=True)
    print(f"📄 Sheets in {output_file_path}: {wb.sheetnames}")

    # Read data
    data_frames = {
        sheet: pd.read_excel(output_file_path, sheet_name=sheet)
        for sheet in component_sheets
    }

    all_techs = sorted(
        set().union(*(df["key_1"].unique() for df in data_frames.values()))
    )

    years = [2025, 2030, 2035, 2040]

    for tech in all_techs:
        data = {}
        for sheet, df in data_frames.items():
            tech_df = df[df["key_1"] == tech].set_index("year").sort_index()
            series = tech_df["value"] / 1000  # Convert to GW
            data[sheet] = series.reindex(years).fillna(0)

        abs_data = pd.DataFrame(data)
        if abs_data.sum().sum() == 0:
            print(f"⚠ No data for technology '{tech}', skipping plots.")
            continue

        # Calculate total additions (sum of all 4 components)
        total_additions = abs_data.sum(axis=1)

        # Calculate each component as percentage of TOTAL additions
        rel_data = pd.DataFrame()
        rel_data["capacity_ext_eusecondary"] = abs_data["capacity_ext_eusecondary"] / total_additions
        rel_data["capacity_ext_stockout"] = abs_data["capacity_ext_stockout"] / total_additions
        rel_data["capacity_ext_euprimary"] = abs_data["capacity_ext_euprimary"] / total_additions

        # Handle division by zero
        rel_data = rel_data.fillna(0)
        rel_data = rel_data.replace([np.inf, -np.inf], 0)

        # === Styling ===
        colors = ["#FDC5B5", "#F99B7D", "#F76C5E"]  # Soft peach to coral
        hatches = ["..", "//", "xx"]
        labels = ["Remanufacturing", "Stock", "Manufacturing"]  # Correct order
        sheet_order = [
            "capacity_ext_eusecondary",  # Remanufacturing (first)
            "capacity_ext_stockout",  # Stock (second)
            "capacity_ext_euprimary",  # Manufacturing (third)
        ]

        # RELATIVE PLOT - Now shows % of total capacity additions
        fig_rel, ax_rel = plt.subplots(figsize=(10, 6))
        bar_container = rel_data[sheet_order].plot(
            kind="bar",
            stacked=True,
            color=colors,
            ax=ax_rel,
            width=0.5,
            edgecolor="black",
            linewidth=0.5,
        )

        for bar_group, hatch in zip(bar_container.containers, hatches):
            for bar in bar_group:
                bar.set_hatch(hatch)

        ax_rel.set_title(
            f"Local Sourcing as % of Total Capacity Additions - {tech}", pad=15
        )
        ax_rel.set_xlabel("Year", labelpad=10)
        ax_rel.set_ylabel("% of Total Capacity Additions", labelpad=10)
        ax_rel.set_xticks(range(len(years)))
        ax_rel.set_xticklabels(years, rotation=45, ha="right")

        # Set y-axis as percentage (max 100% since it's part of total)
        ax_rel.set_ylim(0, 1.0)
        ax_rel.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: '{:.0%}'.format(y)))

        # Add 40% reference line for NZIA benchmark
        benchmark_line = ax_rel.axhline(y=0.4, color='red', linestyle='--', alpha=0.7, linewidth=2)

        # Create combined legend with correct labels
        bar_handles = [container[0] for container in bar_container.containers]
        all_handles = bar_handles + [benchmark_line]
        all_labels = labels + ['NZIA Benchmark']
        ax_rel.legend(all_handles, all_labels, frameon=True, loc="upper right")

        ax_rel.grid(axis="y", alpha=0.3)

        # ABSOLUTE PLOT - Only the 3 local components (exclude imported)
        abs_data_display = abs_data[sheet_order]

        fig_abs, ax_abs = plt.subplots(figsize=(10, 6))
        bar_container_abs = abs_data_display.plot(
            kind="bar",
            stacked=True,
            color=colors,
            ax=ax_abs,
            width=0.5,
            edgecolor="black",
            linewidth=0.5,
        )

        for bar_group, hatch in zip(bar_container_abs.containers, hatches):
            for bar in bar_group:
                bar.set_hatch(hatch)

        ax_abs.set_title(f"Absolute Local Sourcing Capacity - {tech}", pad=15)
        ax_abs.set_ylabel("Capacity (GW)", labelpad=10)
        ax_abs.set_xlabel("Year", labelpad=10)
        ax_abs.set_xticks(range(len(years)))
        ax_abs.set_xticklabels(years, rotation=45, ha="right")
        ax_abs.legend(labels, frameon=True, loc="upper right")  # Only bar labels for absolute plot
        ax_abs.grid(axis="y", alpha=0.3)

        # Save figures
        safe_tech_name = tech.replace(" ", "_").replace("/", "_")
        fig_rel.savefig(
            os.path.join(output_dir, f"relative_composition_{safe_tech_name}.png"),
            dpi=300,
            bbox_inches="tight",
        )
        fig_abs.savefig(
            os.path.join(output_dir, f"absolute_capacity_{safe_tech_name}.png"),
            dpi=300,
            bbox_inches="tight",
        )

        plt.close(fig_rel)
        plt.close(fig_abs)

        print(f"✔ Plots saved for: {tech} in {output_dir}")

        # Debug: Show percentage calculations
        print(f"   Local Sourcing Analysis for {tech}:")
        for year in years:
            if year in total_additions.index and total_additions[year] > 0:
                total = total_additions[year]
                local_total = abs_data_display.loc[year].sum()
                imported = abs_data["capacity_ext_imported"][year]
                local_percentage = (local_total / total) * 100
                print(
                    f"     {year}: Local={local_total:.1f}GW, Imported={imported:.1f}GW, Total={total:.1f}GW, Local%={local_percentage:.1f}%")


def plot_scrap(output_file_path):
    # Extract scenario name from file name
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")

    # Create output folder next to the Excel file
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    # Load scrap data
    sheet_name = "Total_Scrap"
    df = pd.read_excel(output_file_path, sheet_name=sheet_name)

    # Convert from tonnes to megatonnes (Mt)
    df["value"] = df["value"] / 1e6

    # Define the year range
    years = list(range(2024, 2041))

    # Get unique technologies
    techs = df["key_1"].unique()

    for tech in techs:
        tech_df = df[df["key_1"] == tech].set_index("year").sort_index()
        series = tech_df["value"].reindex(years).fillna(0)

        if series.sum() == 0:
            print(f"⚠ No data for technology '{tech}', skipping.")
            continue

        # Plot
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(
            series.index, series.values, color="seagreen", linewidth=2
        )  # smooth line

        ax.set_title(f"Scrap volume – {tech}")
        ax.set_xlabel("Year")
        ax.set_ylabel("Scrap [Mt]")
        ax.set_xlim(2023, 2041)
        ax.set_xticks([2025, 2030, 2035, 2040])  # Only show actual data years
        ax.set_ylim(0, max(series.values) * 1.1 if series.values.any() else 1)
        ax.grid(True, linestyle="--", alpha=0.3)

        plt.tight_layout()
        plot_filename = f"scrap_{tech}.png"
        fig.savefig(os.path.join(output_dir, plot_filename), dpi=300)
        plt.close(fig)

        print(f"✔ Plot saved for: {tech} → {plot_filename}")


def plot_balance_created(output_file_path):
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    # Load data
    df = pd.read_excel(output_file_path, sheet_name="Balance")

    # Filter out Batteries
    df = df[df["key_1"] != "Batteries"]

    # Pivot data
    pivot_df = df.pivot_table(
        index="year", columns="key_1", values="value", aggfunc="sum"
    )
    pivot_df = pivot_df.sort_index()
    pivot_df = pivot_df / 1000  # MWh to GWh
    years = [2025, 2030, 2035, 2040]
    pivot_df = pivot_df.reindex(years).fillna(0)

    # Sort columns consistently
    techs = sorted(pivot_df.columns)
    pivot_df = pivot_df[techs]

    # Generate a beautiful qualitative color palette
    n_techs = len(techs)
    colors = sns.color_palette("Set3", n_colors=n_techs)

    # === Absolute plot ===
    fig_abs, ax_abs = plt.subplots(figsize=(11, 6))
    pivot_df.plot(
        kind="bar",
        stacked=True,
        color=colors,
        ax=ax_abs,
        edgecolor="black",
        linewidth=0.3,
        width=0.6,
    )

    ax_abs.set_title("Created Balance (excl. Batteries)", pad=15)
    ax_abs.set_ylabel("Capacity (GW)")
    ax_abs.set_xlabel("Year")
    ax_abs.set_xticks(range(len(years)))
    ax_abs.set_xticklabels(years, rotation=45, ha="right")
    ax_abs.grid(axis="y", alpha=0.3)
    ax_abs.legend(title="Technology", frameon=True, loc="upper right")  # Legend inside

    fig_abs.tight_layout()
    fig_abs.savefig(
        os.path.join(output_dir, "created_Balance_absolute.png"),
        dpi=300,
        bbox_inches="tight",
    )

    # === Relative plot ===
    rel_df = pivot_df.div(pivot_df.sum(axis=1), axis=0).fillna(0)

    fig_rel, ax_rel = plt.subplots(figsize=(11, 6))
    rel_df.plot(
        kind="bar",
        stacked=True,
        color=colors,
        ax=ax_rel,
        edgecolor="black",
        linewidth=0.3,
        width=0.6,
    )

    ax_rel.set_title("Relative Installed Balance Share (excl. Batteries)", pad=15)
    ax_rel.set_ylabel("Share of Total Capacity")
    ax_rel.set_xlabel("Year")
    ax_rel.set_xticks(range(len(years)))
    ax_rel.set_xticklabels(years, rotation=45, ha="right")
    ax_rel.set_yticklabels(["{:.0%}".format(y) for y in ax_rel.get_yticks()])
    ax_rel.grid(axis="y", alpha=0.3)
    ax_rel.legend(title="Technology", frameon=True, loc="upper right")  # Legend inside

    fig_rel.tight_layout()
    fig_rel.savefig(
        os.path.join(output_dir, "installed_balance_relative.png"),
        dpi=300,
        bbox_inches="tight",
    )

    plt.close(fig_abs)
    plt.close(fig_rel)

    print(f"✔ Installed balance plots saved in {output_dir}")


def lineplot_fuels(output_file_path):
    # Extract scenario name from file name
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")

    # Create output folder next to the Excel file
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    # Load scrap data
    sheet_name = "Commodities_Demand"
    df = pd.read_excel(output_file_path, sheet_name=sheet_name)

    # Define the year range
    years = list(range(2024, 2041))

    # Get unique technologies
    df["key_2"] = df["key_2"].astype(str).str.strip().str.upper()
    fuels = ["PIPED GAS", "LNG"]  # all uppercase now

    for fuel in fuels:
        # Filter for the fuel
        fuel_df = df[df["key_2"] == fuel].copy()
        if fuel_df.empty:
            print(f"⚠ No data found for {fuel}, skipping.")
            continue

        # Convert 'value' column to bcm
        fuel_df["value_bcm"] = mwh_to_bcm(fuel_df["value"])

        # Aggregate by year and reindex to cover all years
        series = fuel_df.groupby("year")["value_bcm"].sum().reindex(years).fillna(0)

        if series.sum() == 0:
            print(f"⚠ No data for fuel '{fuel}', skipping.")
            continue

        # Plot
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.plot(series.index, series.values, color="seagreen", linewidth=2)
        ax.set_title(f"Fuel Demand – {fuel}")
        ax.set_xlabel("Year")
        ax.set_ylabel("Demand (bcm)")
        ax.set_xlim(2023, 2041)
        ax.set_xticks([2025, 2030, 2035, 2040])
        ax.set_ylim(0, max(series.values) * 1.1)
        ax.grid(True, linestyle="--", alpha=0.3)

        plt.tight_layout()
        plot_filename = f"fuel_{fuel.replace(' ', '_').lower()}.png"
        fig.savefig(os.path.join(output_dir, plot_filename), dpi=300)
        plt.close(fig)

        print(f"✔ Plot saved for: {fuel} → {plot_filename}")


def commodities_demand(output_file_path):
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    # Load data
    df = pd.read_excel(output_file_path, sheet_name="Commodities_Demand")

    # Filter out
    exclude = ["Biomass", "Coal", "Lignite", "Hydro", "Nuclear Fuel"]
    df = df[~df["key_2"].isin(exclude)]

    # Pivot data
    pivot_df = df.pivot_table(
        index="year", columns="key_2", values="value", aggfunc="sum"
    )
    pivot_df = pivot_df.sort_index()
    pivot_df = pivot_df
    years = [2025, 2030, 2035, 2040]
    pivot_df = pivot_df.reindex(years).fillna(0)

    # Sort columns consistently
    techs = sorted(pivot_df.columns)
    pivot_df = pivot_df[techs]

    # Generate a beautiful qualitative color palette
    n_techs = len(techs)
    colors = sns.color_palette("Set3", n_colors=n_techs)

    # === Absolute plot ===
    fig_abs, ax_abs = plt.subplots(figsize=(11, 6))
    pivot_df.plot(
        kind="bar",
        stacked=True,
        color=colors,
        ax=ax_abs,
        edgecolor="black",
        linewidth=0.3,
        width=0.6,
    )

    ax_abs.set_title("Fuel Demand", pad=15)
    ax_abs.set_ylabel("unit")
    ax_abs.set_xlabel("Year")
    ax_abs.set_xticks(range(len(years)))
    ax_abs.set_xticklabels(years, rotation=45, ha="right")
    ax_abs.grid(axis="y", alpha=0.3)
    ax_abs.legend(title="Fuel Type", frameon=True, loc="upper right")  # Legend inside

    fig_abs.tight_layout()
    fig_abs.savefig(
        os.path.join(output_dir, "commodities_demand.png"),
        dpi=300,
        bbox_inches="tight",
    )

    plt.close(fig_abs)


def plot_facility_utilization(output_file_path):
    # Extract scenario name and create output directory
    file_name = os.path.basename(output_file_path)
    scenario_name = file_name.replace("result_scenario_", "").replace(".xlsx", "")
    base_dir = os.path.dirname(output_file_path)
    output_dir = os.path.join(base_dir, f"figures_{scenario_name}")
    os.makedirs(output_dir, exist_ok=True)

    # Read data from both sheets
    total_capacity = pd.read_excel(output_file_path, sheet_name="Total Cap Fac")
    used_capacity = pd.read_excel(
        output_file_path, sheet_name="capacity_ext_eusecondary"
    )

    # Define the years we want to plot
    years = [2025, 2030, 2035, 2040]

    # Extract locations and technologies
    locations = total_capacity["key_0"].unique()
    technologies = total_capacity["key_1"].unique()

    # Create a figure for each technology
    for tech in technologies:
        fig, ax = plt.subplots(figsize=(10, 6))

        # Filter data for current technology
        tech_total = total_capacity[total_capacity["key_1"] == tech]
        tech_used = used_capacity[used_capacity["key_1"] == tech]

        # Calculate positions for bars
        n_years = len(years)
        n_locations = len(locations)
        width = 0.35  # width of bars

        # Create position arrays for bars
        indices = np.arange(n_years)

        # Plot bars for each location
        for i, loc in enumerate(locations):
            # Calculate x positions for this location's bars
            x = indices + (i - (n_locations - 1) / 2) * (width + 0.1)

            # Get values for each year
            total_vals = [
                tech_total[(tech_total["key_0"] == loc) & (tech_total["year"] == year)][
                    "value"
                ].values[0]
                / 1000
                if len(
                    tech_total[
                        (tech_total["key_0"] == loc) & (tech_total["year"] == year)
                    ]
                )
                > 0
                else 0
                for year in years
            ]

            used_vals = [
                tech_used[(tech_used["key_0"] == loc) & (tech_used["year"] == year)][
                    "value"
                ].values[0]
                / 1000
                if len(
                    tech_used[(tech_used["key_0"] == loc) & (tech_used["year"] == year)]
                )
                > 0
                else 0
                for year in years
            ]

            # Plot bars with thicker outlines
            ax.bar(
                x,
                total_vals,
                width,
                label=f"Total Capacity {loc}" if i == 0 else "",
                color="lightgray",
                alpha=0.7,
                edgecolor="black",
                linewidth=1.5,
            )
            ax.bar(
                x,
                used_vals,
                width,
                label=f"Used Capacity {loc}" if i == 0 else "",
                color="darkblue",
                alpha=0.7,
                edgecolor="black",
                linewidth=1.5,
            )

        # Customize the plot
        ax.set_xlabel("Year", labelpad=10)
        ax.set_ylabel("Capacity (GW)", labelpad=10)
        ax.set_title(f"Facility Capacity Utilization - {tech}", pad=15)
        ax.set_xticks(indices)
        ax.set_xticklabels(years, rotation=45, ha="right")

        # Add legend
        handles = []
        labels = []
        for loc in locations:
            handles.extend(
                [
                    plt.Rectangle(
                        (0, 0),
                        1,
                        1,
                        color="lightgray",
                        alpha=0.7,
                        edgecolor="black",
                        linewidth=1.5,
                    ),
                    plt.Rectangle(
                        (0, 0),
                        1,
                        1,
                        color="darkblue",
                        alpha=0.7,
                        edgecolor="black",
                        linewidth=1.5,
                    ),
                ]
            )
            labels.extend([f"Total Capacity {loc}", f"Used Capacity {loc}"])

        ax.legend(
            handles, labels, frameon=True, loc="upper left", bbox_to_anchor=(1, 1)
        )

        # Add grid for better readability
        ax.grid(axis="y", alpha=0.3)

        # Adjust layout to prevent legend cutoff
        plt.tight_layout()

        # Save the figure
        safe_tech_name = tech.replace(" ", "_").replace("/", "_")
        plt.savefig(
            os.path.join(output_dir, f"facility_utilization_{safe_tech_name}.png"),
            dpi=300,
            bbox_inches="tight",
        )
        plt.close(fig)

    print(f"✅ Facility utilization plots saved in {output_dir}")


# Example: Call this after saving Excel
# write_carryovers_to_excel(..., output_file_path)
# plot_capacity_decomposition_by_technology(output_file_path)
def plot_all_scenarios(base_dir):  # TODO re-add if needed
    # Find all Excel files in the base_dir and its subfolders
    excel_files = glob.glob(
        os.path.join(base_dir, "**", "result_scenario_*.xlsx"), recursive=True
    )

    for file in excel_files:
        print(f"📊 Processing: {file}")
        expected_sheets = [
            "Installed_Capacity_Q_s",
            "Existing_Stock_Q_stock",
            "capacity_dec_start",
            "Total Cap Sec",
            "Total Cap Fac",
            "CO2_emissions",
            "Total_Cost",
            "Total_Scrap",
            "Pricereduction",
            "Balance",
            "Commodities_Demand",
            "capacity_ext_imported",
            "capacity_ext_stockout",
            "capacity_ext_euprimary",
            "capacity_ext_eusecondary",
            "capacity_ext_stock_imported",
            "newly_added_capacity",
            "capacity_facility_eusecondary",
        ]

        wait_for_excel_sheets(file, expected_sheets)
        plot_capacity_decomposition_by_technology(file)


def wait_for_excel_sheets(path, expected_sheets, timeout=60):  # TODO re-add if needed
    """Wait until the expected sheets exist in the Excel file."""
    start = time.time()
    while time.time() - start < timeout:
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            if all(sheet in sheets for sheet in expected_sheets):
                return True
        except Exception:
            pass
        time.sleep(0.5)
    raise TimeoutError(f"Expected sheets not found in {path} after {timeout} seconds.")


# plot_nzia_benchmark("result/urbs-20250611T1324/result_scenario_base.xlsx")
# plot_all_scenarios("result")
# plot_scrap("result/urbs-20250520T1651/result_scenario_base.xlsx")
# plot_installed_capacity("result/urbs-20250520T1651/result_scenario_base.xlsx")
# plot_commodities_demand("result/urbs-20250604T1424/result_scenario_base.xlsx")
# commodities_demand("result/urbs-20250716T1015/result_scenario_base.xlsx")
# lineplot_fuels("result/urbs-20250716T1015/result_scenario_base.xlsx")

# plot_balance_created("result/urbs-20250604T1538/result_scenario_base.xlsx")
# plot_facility_utilization("result/urbs-20250716T1015/result_scenario_base.xlsx")
plot_nzia_benchmark(r"C:\Users\maxoi\OneDrive\Desktop\results_crm_paper\LR7\result_scenario_extremely_high.xlsx")
